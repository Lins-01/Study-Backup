# 处理excel表格的包openpyxl，需要下载
import openpyxl as xl

# 画表格功能 ： 再导入其中chart module中的两个类(名字有大写字母)
# Reference是用来得到一个范围的数据的（两个坐标，即四个值表示范围）
from openpyxl.chart import BarChart, Reference

# 封装成函数，可以传入不同的文件名
# 删去多余的行


def process_workbook(filename):

    # 获取到文件 
    wb = xl.load_workbook(filename)
    # 获取到对应的sheet页，要用大写
    sheet = wb['Sheet1']

    # 需要+1 ： range是左闭右开
    for row in range(2, sheet.max_row+1):

        cell = sheet.cell(row, 3)

        corrected_price = cell.value*0.9

        # 拿到需要存储的位置对象
        corrected_price_cell = sheet.cell(row, 4)
        # 修改值为需要的
        corrected_price_cell.value = corrected_price

    # 获取到需要操作的区域数据，这里只操作上面生成的这列
    values = Reference(sheet,
                       4,
                       2,
                       4,
                       sheet.max_row)

    # 实例化表格对象
    chart = BarChart()
    # 给表格中添加数据
    chart.add_data(values)
    # sheet页中添加表格 ,第二个参数是画表格的坐标
    sheet.add_chart(chart, 'e2')

    # 把整个文件对象保存。
    # 存为一个新文件
    wb.save(filename)
