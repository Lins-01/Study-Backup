# 处理excel表格的包openpyxl，需要下载
import openpyxl as xl

# 画表格功能 ： 再导入其中chart module中的两个类(名字有大写字母)
# Reference是用来得到一个范围的数据的（两个坐标，即四个值表示范围）
from openpyxl.chart import BarChart, Reference

import os

cwd = os.getcwd()  # Get the current working directory (cwd)
files = os.listdir(cwd)  # Get all the files in that directory
print("Files in %r: %s" % (cwd, files))

# 获取到文件
wb = xl.load_workbook('transactions.xlsx')
# 获取到对应的sheet页，要用大写
sheet = wb['Sheet1']

# 获取第一1第一列的方式，a1指：1行a列
cell = sheet['a1']
print(cell)
print(cell.value)
cell = sheet.cell(1, 1)
print(cell)
print(cell.value)

# 打印sheet的行数和列数
print(sheet.max_row)
print(sheet.max_column)

# 需要+1 ： range是左闭右开
for row in range(2, sheet.max_row+1):
    print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value*0.9
    print(f'corrected_price:  {corrected_price}')
    # 拿到需要存储的位置对象
    corrected_price_cell = sheet.cell(row, 4)
    # 修改值为需要的
    corrected_price_cell.value = corrected_price



# 获取到需要操作的区域数据，这里只操作上面生成的这列
values = Reference(sheet,
                   4,
                   2,
                   4,
                   sheet.max_row)

print(values)

# 实例化表格对象
chart = BarChart()
# 给表格中添加数据
chart.add_data(values)
# sheet页中添加表格 ,第二个参数是画表格的坐标
sheet.add_chart(chart,'e2')


# 把整个文件对象保存。
# 存为一个新文件
wb.save('transactions2.xlsx')
